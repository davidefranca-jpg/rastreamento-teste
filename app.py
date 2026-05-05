# -*- coding: utf-8 -*-
"""
SISTEMA DE RASTREAMENTO POR GRUPOS - VERSAO PRO INICIAL
Caminho recomendado:
    C:\IA\Projeto\app_rastreamento_grupos_pro.py

Objetivo:
- Separar o sistema em 4 paginas/perfis:
  1) Cliente
  2) Admin
  3) Rastreamento
  4) Expedicao

Foco desta versao:
- Pagina Rastreamento pronta para operacao diaria.
- Outras paginas criadas como base estrutural.
- Upload diario do relatorio geral por Admin/Gestor.
- Controle por grupo, parceiro, UF, cliente, SLA, atraso e ocorrencia.
- Cadastro de grupos com 2 pessoas.
- Download individual ou do grupo por senha.
- Upload de retorno dos grupos com status/ocorrencia.
- Historico anual em SQLite, sem apagar o passado.
- Pesquisa por NF mostrando status, grupo e parceiro.

Como rodar:
    cd C:\IA\Projeto
    .\venv\Scripts\activate
    pip install openpyxl
    python C:\IA\Projeto\app_rastreamento_grupos_pro.py

Acessar:
    http://127.0.0.1:8092

Login inicial:
    usuario: admin
    senha: Admin@123
"""

import os
import re
import io
import csv
import hmac
import json
import uuid
import html
import time
import sqlite3
import hashlib
import secrets
import traceback
from http.cookies import SimpleCookie
from datetime import datetime, date, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, urlencode

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:
    print("ERRO: biblioteca openpyxl nao instalada.")
    print(r"Rode: cd C:\IA\Projeto")
    print(r"Depois: .\venv\Scripts\activate")
    print("Depois: pip install openpyxl")
    raise

HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", "8092"))

if os.name == "nt":
    APP_DIR = os.environ.get("APP_DIR", r"C:\IA\Projeto")
else:
    APP_DIR = os.environ.get("APP_DIR", os.path.dirname(os.path.abspath(__file__)))

DATA_DIR = os.path.join(APP_DIR, "dados_rastreamento")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
DOWNLOAD_DIR = os.path.join(DATA_DIR, "downloads")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "rastreamento_grupos.db")
SESSIONS_FILE = os.path.join(DATA_DIR, "sessoes_rastreamento.json")
SESSION_HOURS = 24

ADMIN_LOGIN = os.environ.get("ADMIN_LOGIN", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Admin@123")

STATUS_OPTIONS = [
    "Entregue no prazo",
    "Entregue atrasado",
    "Em aberto",
    "Atrasado com ocorrência",
    "Cancelado",
]

CSV_SEPARATOR = ";"

COL_ALIASES = {
    "nf": ["Nota Fiscal", "NF", "Nota", "Nf"],
    "data_inicial": ["Data Inicial", "Data", "Emissao", "Data Emissao"],
    "uf_origem": ["UF Remetente", "UF Origem"],
    "tipo": ["Tipo"],
    "entrega": ["Entrega"],
    "link": ["Link Comprovante", "Comprovante"],
    "destinatario": ["Destinatário", "Destinatario"],
    "cidade_destino": ["CidadeDestinatário", "Cidade Destinatario", "CidadeDestino"],
    "uf_destino": ["UFDest.", "UF Destino", "UFDest"],
    "data_entrega": ["Data deEntrega", "Data de Entrega", "Data Entrega"],
    "ocorrencia": ["Ocorrência", "Ocorrencia"],
    "filial": ["Filial"],
    "sla": ["SLA acordado", "SLA", "Pazos", "Prazo"],
    "parceiro": ["Parceiros", "Parceiro", "PARCEIRO"],
    "sla_justificado": ["Sla Justificado", "SLA Justificado"],
    "uf_parceiro": ["UF doa parceiros", "UF dos parceiros", "UF parceiro", "Nome_UF"],
    "cliente_codigo": ["Codigo Cliente", "Código Cliente", "Codigo Fatura Para", "Cod Cliente"],
    "cliente_nome": ["Nome do cliente", "Nome Cliente", "Faturar Para", "Cliente"],
}

# Fallback para a planilha enviada no projeto.
FALLBACK_COLS_1_BASED = {
    "nf": 1,
    "data_inicial": 3,
    "uf_origem": 7,
    "tipo": 9,
    "entrega": 10,
    "link": 11,
    "destinatario": 13,
    "cidade_destino": 14,
    "uf_destino": 15,
    "data_entrega": 18,
    "ocorrencia": 19,
    "filial": 25,
    "sla": 28,
    "parceiro": 29,
    "sla_justificado": 30,
    "uf_parceiro": 31,
    "cliente_codigo": 32,
    "cliente_nome": 33,
}

# Colunas fixas confirmadas na aba Rastreamento:
# AB = prazo/SLA, AC = parceiro/base, AF = código cliente, AG = nome cliente.
RASTREAMENTO_COLS_FIXAS_1_BASED = {
    "sla": 28,
    "parceiro": 29,
    "cliente_codigo": 32,
    "cliente_nome": 33,
}

MAX_QUERY_ROWS = int(os.environ.get("MAX_QUERY_ROWS", "200000"))

SESSIONS = {}


def norm_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return " ".join(s.split())


def norm_key(v):
    s = norm_text(v).upper()
    trans = str.maketrans("ÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ", "AAAAEEEIIIOOOOUUUC")
    return s.translate(trans)


def safe_filename(v):
    s = norm_text(v) or "arquivo"
    s = re.sub(r"[^A-Za-z0-9_.-]+", "_", s)
    return s[:120]


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return date(1899, 12, 30) + timedelta(days=int(v))
        except Exception:
            return None
    s = str(v).strip()
    if re.match(r"^\d{2}:\d{2}", s):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    return None


def fmt_date(d):
    if isinstance(d, str):
        d2 = to_date(d)
        return d2.strftime("%d/%m/%Y") if d2 else d
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


def db_date(d):
    if not d:
        return ""
    if isinstance(d, str):
        d2 = to_date(d)
        return d2.isoformat() if d2 else ""
    return d.isoformat()


def to_int(v, default=0):
    if v is None or str(v).strip() == "":
        return default
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return int(float(m.group(0))) if m else default


def pct(a, b):
    return 0.0 if not b else round((a / b) * 100, 2)


def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000)
    return salt + "$" + digest.hex()


def check_password(password, stored_hash):
    try:
        salt, digest = stored_hash.split("$", 1)
        test = hash_password(password, salt).split("$", 1)[1]
        return hmac.compare_digest(test, digest)
    except Exception:
        return False


def connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = connect()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        login TEXT UNIQUE NOT NULL,
        senha_hash TEXT NOT NULL,
        nome TEXT NOT NULL,
        perfil TEXT NOT NULL,
        grupo_id INTEGER,
        cliente_codigo TEXT DEFAULT '',
        ativo INTEGER DEFAULT 1,
        trocar_senha INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS grupos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL,
        pessoa1_login TEXT DEFAULT '',
        pessoa2_login TEXT DEFAULT '',
        max_diario INTEGER DEFAULT 0,
        ativo INTEGER DEFAULT 1,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS perfis (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL,
        descricao TEXT DEFAULT '',
        ativo INTEGER DEFAULT 1,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS perfil_permissoes (
        perfil_nome TEXT NOT NULL,
        permissao TEXT NOT NULL,
        PRIMARY KEY (perfil_nome, permissao)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS uploads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT NOT NULL,
        filename TEXT,
        path TEXT,
        usuario TEXT,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS demandas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        upload_id INTEGER,
        data_base TEXT,
        ano INTEGER,
        mes INTEGER,
        nf TEXT,
        cliente_codigo TEXT,
        cliente_nome TEXT,
        parceiro TEXT,
        uf_parceiro TEXT,
        estado TEXT,
        cidade_destino TEXT,
        uf_origem TEXT,
        uf_destino TEXT,
        filial TEXT,
        tipo TEXT,
        entrega_original TEXT,
        ocorrencia_original TEXT,
        link_comprovante TEXT,
        data_inicial TEXT,
        data_prevista TEXT,
        data_entrega TEXT,
        sla INTEGER,
        status_sistema TEXT,
        status_grupo TEXT DEFAULT '',
        ocorrencia_grupo TEXT DEFAULT '',
        grupo_id INTEGER,
        responsavel_login TEXT DEFAULT '',
        dias_atraso_sla INTEGER DEFAULT 0,
        atraso_atualizacao_dias INTEGER DEFAULT 0,
        cancelado INTEGER DEFAULT 0,
        finalizado INTEGER DEFAULT 0,
        updated_by TEXT DEFAULT '',
        updated_at TEXT DEFAULT '',
        active INTEGER DEFAULT 1,
        raw_json TEXT DEFAULT ''
    )
    """)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_nf ON demandas(nf)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_active ON demandas(active)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_grupo ON demandas(grupo_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_cliente ON demandas(cliente_codigo)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_parceiro ON demandas(parceiro)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_demandas_ano_mes ON demandas(ano, mes)")

    perfis_padrao = {
        "admin": ["admin", "acompanhamento", "rastreamento", "chamado_parceiro", "cliente", "expedicao", "usuarios", "grupos", "upload_geral", "redistribuir"],
        "gestor": ["acompanhamento", "rastreamento", "chamado_parceiro", "upload_geral", "redistribuir"],
        "rastreamento": ["acompanhamento", "rastreamento"],
        "cliente": ["cliente"],
        "expedicao": ["expedicao"],
    }
    for perfil_nome, permissoes in perfis_padrao.items():
        cur.execute("INSERT OR IGNORE INTO perfis(nome, descricao, ativo, created_at) VALUES(?,?,1,?)", (perfil_nome, "Perfil padrão do sistema", datetime.now().isoformat()))
        for permissao in permissoes:
            cur.execute("INSERT OR IGNORE INTO perfil_permissoes(perfil_nome, permissao) VALUES(?,?)", (perfil_nome, permissao))

    cur.execute("SELECT id FROM users WHERE login=?", (ADMIN_LOGIN,))
    if not cur.fetchone():
        cur.execute("""
            INSERT INTO users(login, senha_hash, nome, perfil, grupo_id, cliente_codigo, ativo, trocar_senha, created_at)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (ADMIN_LOGIN, hash_password(ADMIN_PASSWORD), "Administrador", "admin", None, "", 1, 0, datetime.now().isoformat()))

    conn.commit()
    conn.close()


def row_to_dict(row):
    return dict(row) if row else None


def load_sessions():
    global SESSIONS
    if not os.path.exists(SESSIONS_FILE):
        SESSIONS = {}
        return
    try:
        with open(SESSIONS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        now = time.time()
        SESSIONS = {k: v for k, v in data.items() if now - v.get("created", 0) <= SESSION_HOURS * 3600}
    except Exception:
        SESSIONS = {}


def save_sessions():
    try:
        with open(SESSIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(SESSIONS, f, ensure_ascii=False, indent=2)
    except Exception:
        traceback.print_exc()


def get_session_user(handler):
    cookie_header = handler.headers.get("Cookie", "")
    if not cookie_header:
        return None
    c = SimpleCookie()
    try:
        c.load(cookie_header)
    except Exception:
        return None
    sid = c.get("sid")
    if not sid:
        return None
    sess = SESSIONS.get(sid.value)
    if not sess:
        return None
    if time.time() - sess.get("created", 0) > SESSION_HOURS * 3600:
        SESSIONS.pop(sid.value, None)
        save_sessions()
        return None
    conn = connect()
    u = conn.execute("SELECT * FROM users WHERE login=? AND ativo=1", (sess.get("login"),)).fetchone()
    conn.close()
    return row_to_dict(u)


def create_session(handler, login):
    sid = secrets.token_urlsafe(32)
    SESSIONS[sid] = {"login": login, "created": time.time()}
    save_sessions()
    handler.send_header("Set-Cookie", f"sid={sid}; Path=/; HttpOnly; SameSite=Lax")


def clear_session(handler):
    cookie_header = handler.headers.get("Cookie", "")
    if cookie_header:
        try:
            c = SimpleCookie()
            c.load(cookie_header)
            sid = c.get("sid")
            if sid:
                SESSIONS.pop(sid.value, None)
                save_sessions()
        except Exception:
            pass
    handler.send_header("Set-Cookie", "sid=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")


def read_form_urlencoded(handler):
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length).decode("utf-8", "ignore")
    data = parse_qs(raw)
    return {k: v[0] if v else "" for k, v in data.items()}


def parse_multipart(handler):
    ctype = handler.headers.get("Content-Type", "")
    m = re.search("boundary=(.*)", ctype)
    if not m:
        raise ValueError("Upload invalido: boundary nao encontrado.")
    boundary = ("--" + m.group(1)).encode()
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    parts = body.split(boundary)

    fields = {}
    files = {}
    for part in parts:
        if b"\r\n\r\n" not in part:
            continue
        head, content = part.split(b"\r\n\r\n", 1)
        content = content.rsplit(b"\r\n", 1)[0]
        name_m = re.search(rb'name="([^"]+)"', head)
        if not name_m:
            continue
        name = name_m.group(1).decode("utf-8", "ignore")
        filename_m = re.search(rb'filename="([^"]*)"', head)
        if filename_m:
            filename = filename_m.group(1).decode("utf-8", "ignore") or "upload.xlsx"
            files[name] = {"filename": filename, "content": content}
        else:
            fields[name] = content.decode("utf-8", "ignore")
    return fields, files


def has_permission(user, allowed):
    if not user:
        return False
    perfil = norm_text(user.get("perfil"))
    if perfil in allowed:
        return True
    try:
        conn = connect()
        marks = ",".join("?" for _ in allowed)
        row = conn.execute(f"SELECT 1 FROM perfil_permissoes WHERE perfil_nome=? AND permissao IN ({marks}) LIMIT 1", [perfil] + list(allowed)).fetchone()
        conn.close()
        return bool(row)
    except Exception:
        return False


PERMISSOES_SISTEMA = [
    ("admin", "Admin"),
    ("acompanhamento", "Acompanhamento"),
    ("rastreamento", "Rastreamento"),
    ("chamado_parceiro", "Chamado Parceiro"),
    ("cliente", "Cliente"),
    ("expedicao", "Expedição"),
    ("usuarios", "Usuários"),
    ("grupos", "Grupos"),
    ("upload_geral", "Upload relatório geral"),
    ("redistribuir", "Redistribuir demanda"),
]


def perfil_options(selected=""):
    conn = connect()
    rows = conn.execute("SELECT nome FROM perfis WHERE ativo=1 ORDER BY nome").fetchall()
    conn.close()
    out = []
    for r in rows:
        nome = norm_text(r["nome"])
        sel = " selected" if nome == selected else ""
        out.append(f"<option value='{html_escape(nome)}'{sel}>{html_escape(nome)}</option>")
    return "".join(out)


def render_permissoes_checkboxes(marcadas=None):
    marcadas = set(marcadas or [])
    itens = []
    for codigo, label in PERMISSOES_SISTEMA:
        checked = " checked" if codigo in marcadas else ""
        itens.append(f"<label class='check_item'><input type='checkbox' name='perm_{html_escape(codigo)}' value='1'{checked}> {html_escape(label)}</label>")
    return "".join(itens)


def render_perfis_admin():
    conn = connect()
    perfis = conn.execute("SELECT * FROM perfis ORDER BY nome").fetchall()
    linhas = []
    for p in perfis:
        perms = [r["permissao"] for r in conn.execute("SELECT permissao FROM perfil_permissoes WHERE perfil_nome=? ORDER BY permissao", (p["nome"],)).fetchall()]
        linhas.append(f"""
        <tr>
            <td>{html_escape(p['nome'])}</td>
            <td>{html_escape(p['descricao'])}</td>
            <td>{html_escape(', '.join(perms))}</td>
        </tr>
        """)
    conn.close()
    return f"""
    <section class='card'>
        <h2>Criar perfil e permissões</h2>
        <p class='note'>Crie perfis personalizados e marque quais páginas/funções cada perfil poderá acessar.</p>
        <form method='post' action='/criar_perfil' class='perfil_form'>
            <div><label>Nome do perfil</label><input name='nome' required placeholder='ex: supervisor'></div>
            <div><label>Descrição</label><input name='descricao' placeholder='Descrição do perfil'></div>
            <div class='permissoes_box'>{render_permissoes_checkboxes()}</div>
            <button>Criar/Atualizar perfil</button>
        </form>
        <div class='tablebox small'><table><tr><th>Perfil</th><th>Descrição</th><th>Permissões</th></tr>{''.join(linhas)}</table></div>
    </section>
    """


def html_escape(v):
    return html.escape(norm_text(v))


def read_holidays(wb):
    holidays = set()
    for sheet in ("Feriado - final de semana", "Listas"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                d = to_date(cell)
                if d:
                    holidays.add(d)
    return holidays


def is_business_day(d, holidays):
    return d.weekday() < 5 and d not in holidays


def add_business_days(start, days, holidays):
    if not start:
        return None
    days = max(0, to_int(days, 0))
    d = start
    added = 0
    while added < days:
        d += timedelta(days=1)
        if is_business_day(d, holidays):
            added += 1
    return d


def business_days_between(start, end, holidays):
    if not start or not end or end <= start:
        return 0
    d = start + timedelta(days=1)
    total = 0
    while d <= end:
        if is_business_day(d, holidays):
            total += 1
        d += timedelta(days=1)
    return total


def find_header_map(ws):
    header = [norm_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    normalized = {norm_key(v): i + 1 for i, v in enumerate(header) if norm_text(v)}
    colmap = {}
    for key, names in COL_ALIASES.items():
        found = None
        for name in names:
            nk = norm_key(name)
            if nk in normalized:
                found = normalized[nk]
                break
        colmap[key] = found or FALLBACK_COLS_1_BASED.get(key)

    # Regra fixa da planilha do projeto: não deixar cabeçalho parecido deslocar colunas críticas.
    colmap.update(RASTREAMENTO_COLS_FIXAS_1_BASED)
    return colmap


def cell(row, col_1_based):
    if not col_1_based:
        return None
    idx = col_1_based - 1
    return row[idx] if idx < len(row) else None


def classify_demand(r, holidays, hoje):
    data_inicial = r.get("data_inicial")
    data_entrega = r.get("data_entrega")
    sla = r.get("sla") or 0
    data_prevista = add_business_days(data_inicial, sla, holidays)

    entrega_key = norm_key(r.get("entrega_original"))
    occ_key = norm_key(r.get("ocorrencia_original"))
    sla_just_key = norm_key(r.get("sla_justificado"))

    cancelado = 1 if "CANCEL" in entrega_key or "CANCEL" in occ_key else 0
    entregue = bool(data_entrega) or "ENTREG" in entrega_key or "FINALIZ" in entrega_key
    ocorrencia = bool(norm_text(r.get("ocorrencia_original"))) and "DOCUMENTO EMITIDO" not in occ_key
    justificado = "JUSTIFICADO" in sla_just_key

    if cancelado:
        status = "Cancelado"
        finalizado = 1
    elif entregue and data_entrega and data_prevista and data_entrega <= data_prevista:
        status = "Entregue no prazo"
        finalizado = 1
    elif entregue:
        status = "Entregue atrasado"
        finalizado = 1
    elif data_prevista and hoje <= data_prevista:
        status = "Em aberto no prazo"
        finalizado = 0
    else:
        status = "Em aberto com atraso"
        finalizado = 0

    if status == "Em aberto com atraso" and ocorrencia:
        status = "Em aberto com ocorrência"

    dias_atraso = 0
    if data_prevista:
        fim = data_entrega if data_entrega else hoje
        if fim > data_prevista:
            dias_atraso = business_days_between(data_prevista, fim, holidays)

    atraso_atualizacao = 0
    if data_prevista and data_entrega:
        limite_atualizacao = data_prevista + timedelta(days=1)
        if data_entrega > limite_atualizacao:
            atraso_atualizacao = business_days_between(limite_atualizacao, data_entrega, holidays)

    return status, data_prevista, dias_atraso, atraso_atualizacao, cancelado, finalizado, justificado


def read_rastreamento_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Rastreamento" not in wb.sheetnames:
        raise ValueError('A planilha precisa ter a aba "Rastreamento".')
    ws = wb["Rastreamento"]
    colmap = find_header_map(ws)
    holidays = read_holidays(wb)
    hoje = date.today()
    rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nf = norm_text(cell(row, colmap["nf"]))
        if not nf:
            continue

        r = {
            "linha_excel": row_idx,
            "nf": nf,
            "data_inicial": to_date(cell(row, colmap["data_inicial"])),
            "uf_origem": norm_text(cell(row, colmap["uf_origem"])),
            "tipo": norm_text(cell(row, colmap["tipo"])),
            "entrega_original": norm_text(cell(row, colmap["entrega"])),
            "link_comprovante": norm_text(cell(row, colmap["link"])),
            "destinatario": norm_text(cell(row, colmap["destinatario"])),
            "cidade_destino": norm_text(cell(row, colmap["cidade_destino"])),
            "uf_destino": norm_text(cell(row, colmap["uf_destino"])),
            "data_entrega": to_date(cell(row, colmap["data_entrega"])),
            "ocorrencia_original": norm_text(cell(row, colmap["ocorrencia"])),
            "filial": norm_text(cell(row, colmap["filial"])),
            "sla": to_int(cell(row, colmap["sla"]), 0),
            "parceiro": norm_text(cell(row, colmap["parceiro"])) or "Parceiro nao informado",
            "sla_justificado": norm_text(cell(row, colmap["sla_justificado"])),
            "uf_parceiro": norm_text(cell(row, colmap["uf_parceiro"])),
            "cliente_codigo": norm_text(cell(row, colmap["cliente_codigo"])),
            "cliente_nome": norm_text(cell(row, colmap["cliente_nome"])),
        }
        r["estado"] = r["uf_parceiro"] or r["uf_destino"] or r["uf_origem"]
        status, prevista, dias_atraso, atraso_atualizacao, cancelado, finalizado, justificado = classify_demand(r, holidays, hoje)
        r["data_prevista"] = prevista
        r["status_sistema"] = status
        r["dias_atraso_sla"] = dias_atraso
        r["atraso_atualizacao_dias"] = atraso_atualizacao
        r["cancelado"] = cancelado
        r["finalizado"] = finalizado
        r["justificado"] = justificado
        rows.append(r)
    return rows


def get_group_by_partner(conn, parceiro, estado):
    parceiro_key = norm_key(parceiro)
    estado_key = norm_key(estado)
    row = conn.execute("""
        SELECT grupo_id FROM demandas
        WHERE active=1
          AND UPPER(parceiro)=UPPER(?)
          AND UPPER(estado)=UPPER(?)
          AND grupo_id IS NOT NULL
        ORDER BY id DESC LIMIT 1
    """, (parceiro, estado)).fetchone()
    if row:
        return row["grupo_id"]
    row = conn.execute("""
        SELECT grupo_id FROM demandas
        WHERE active=1
          AND UPPER(parceiro)=UPPER(?)
          AND grupo_id IS NOT NULL
        ORDER BY id DESC LIMIT 1
    """, (parceiro,)).fetchone()
    return row["grupo_id"] if row else None


def least_loaded_group(conn):
    groups = conn.execute("SELECT id FROM grupos WHERE ativo=1 ORDER BY id").fetchall()
    if not groups:
        return None
    loads = []
    for g in groups:
        c = conn.execute("SELECT COUNT(*) qtd FROM demandas WHERE active=1 AND finalizado=0 AND grupo_id=?", (g["id"],)).fetchone()["qtd"]
        loads.append((c, g["id"]))
    loads.sort()
    return loads[0][1]


def choose_group(conn, parceiro, estado):
    existing = get_group_by_partner(conn, parceiro, estado)
    if existing:
        return existing
    return least_loaded_group(conn)


def membros_do_grupo(conn, grupo_id):
    g = conn.execute("SELECT pessoa1_login, pessoa2_login FROM grupos WHERE id=?", (grupo_id,)).fetchone()
    if not g:
        return []
    return [x for x in (norm_text(g["pessoa1_login"]), norm_text(g["pessoa2_login"])) if x]


def alternar_responsaveis_grupo(conn, grupo_id):
    pessoas = membros_do_grupo(conn, grupo_id)
    if not pessoas:
        return 0
    demandas = conn.execute("""
        SELECT id FROM demandas
        WHERE active=1 AND grupo_id=? AND finalizado=0
        ORDER BY dias_atraso_sla DESC, data_prevista ASC, id ASC
    """, (grupo_id,)).fetchall()
    atualizados = 0
    for idx, d in enumerate(demandas):
        resp = pessoas[idx % len(pessoas)]
        conn.execute("UPDATE demandas SET responsavel_login=? WHERE id=?", (resp, d["id"]))
        atualizados += 1
    return atualizados


def redistribuir_demandas_ativas(force=True):
    """
    Distribui SOMENTE demandas em aberto para grupos/pessoas.
    Fechado, entregue e cancelado ficam fora da fila operacional.
    """
    conn = connect()
    try:
        grupos = [dict(g) for g in conn.execute("SELECT * FROM grupos WHERE ativo=1 ORDER BY id").fetchall()]
        if not grupos:
            return 0, 0, 0

        if force:
            conn.execute("UPDATE demandas SET grupo_id=NULL, responsavel_login='' WHERE active=1")

        cargas = {g["id"]: conn.execute("SELECT COUNT(*) qtd FROM demandas WHERE active=1 AND finalizado=0 AND grupo_id=?", (g["id"],)).fetchone()["qtd"] for g in grupos}
        parceiros = conn.execute("""
            SELECT parceiro, COUNT(*) qtd
            FROM demandas
            WHERE active=1 AND finalizado=0 AND (grupo_id IS NULL OR grupo_id='' OR ?=1)
            GROUP BY parceiro
            ORDER BY qtd DESC, parceiro
        """, (1 if force else 0,)).fetchall()

        parceiros_alocados = 0
        demandas_alocadas = 0
        for p in parceiros:
            parceiro = p["parceiro"] or "Parceiro nao informado"
            qtd = int(p["qtd"] or 0)
            gid = min(cargas.items(), key=lambda x: (x[1], x[0]))[0]
            cur = conn.execute("UPDATE demandas SET grupo_id=? WHERE active=1 AND finalizado=0 AND parceiro=?", (gid, parceiro))
            cargas[gid] += qtd
            parceiros_alocados += 1
            demandas_alocadas += cur.rowcount

        resp_atualizados = 0
        for g in grupos:
            resp_atualizados += alternar_responsaveis_grupo(conn, g["id"])
            for login in (norm_text(g.get("pessoa1_login")), norm_text(g.get("pessoa2_login"))):
                if login:
                    conn.execute("UPDATE users SET grupo_id=? WHERE login=?", (g["id"], login))

        conn.commit()
        return demandas_alocadas, parceiros_alocados, resp_atualizados
    finally:
        conn.close()


def save_upload_record(conn, tipo, filename, path, usuario):
    cur = conn.execute("""
        INSERT INTO uploads(tipo, filename, path, usuario, created_at)
        VALUES(?,?,?,?,?)
    """, (tipo, filename, path, usuario, datetime.now().isoformat()))
    return cur.lastrowid


def import_master_report(path, filename, usuario, ano_controle=None, mes_controle=None):
    rows = read_rastreamento_workbook(path)
    conn = connect()
    try:
        upload_id = save_upload_record(conn, "RELATORIO_GERAL", filename, path, usuario)
        hoje = date.today()
        ano_controle = int(ano_controle or hoje.year)
        mes_controle = int(mes_controle or hoje.month)
        if mes_controle < 1 or mes_controle > 12:
            mes_controle = hoje.month
        # Mantem historico antigo e troca somente a base ativa.
        conn.execute("UPDATE demandas SET active=0 WHERE active=1")

        for r in rows:
            grupo_id = None if r.get("finalizado") else choose_group(conn, r["parceiro"], r["estado"])
            conn.execute("""
                INSERT INTO demandas(
                    upload_id, data_base, ano, mes, nf, cliente_codigo, cliente_nome, parceiro, uf_parceiro,
                    estado, cidade_destino, uf_origem, uf_destino, filial, tipo, entrega_original,
                    ocorrencia_original, link_comprovante, data_inicial, data_prevista, data_entrega, sla,
                    status_sistema, grupo_id, dias_atraso_sla, atraso_atualizacao_dias, cancelado, finalizado,
                    raw_json, active
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
            """, (
                upload_id, hoje.isoformat(), ano_controle, mes_controle, r["nf"], r["cliente_codigo"], r["cliente_nome"],
                r["parceiro"], r["uf_parceiro"], r["estado"], r["cidade_destino"], r["uf_origem"], r["uf_destino"],
                r["filial"], r["tipo"], r["entrega_original"], r["ocorrencia_original"], r["link_comprovante"],
                db_date(r["data_inicial"]), db_date(r["data_prevista"]), db_date(r["data_entrega"]), r["sla"],
                r["status_sistema"], grupo_id, r["dias_atraso_sla"], r["atraso_atualizacao_dias"],
                r["cancelado"], r["finalizado"], json.dumps(r, ensure_ascii=False, default=str)
            ))

        # Auto cria usuarios cliente por codigo, sem resetar senha.
        clientes = {}
        for r in rows:
            if r["cliente_codigo"]:
                clientes[r["cliente_codigo"]] = r["cliente_nome"]
        for cod, nome in clientes.items():
            login = "cliente_" + re.sub(r"[^A-Za-z0-9_.-]+", "_", cod)
            exists = conn.execute("SELECT id FROM users WHERE login=?", (login,)).fetchone()
            if not exists:
                conn.execute("""
                    INSERT INTO users(login, senha_hash, nome, perfil, grupo_id, cliente_codigo, ativo, trocar_senha, created_at)
                    VALUES(?,?,?,?,?,?,?,?,?)
                """, (login, hash_password(f"{cod}@123"), nome or login, "cliente", None, cod, 1, 1, datetime.now().isoformat()))

        conn.commit()

        # Após importar, recalcula a fila operacional somente com demandas em aberto.
        try:
            redistribuir_demandas_ativas(force=True)
        except Exception:
            traceback.print_exc()
        return len(rows), upload_id
    finally:
        conn.close()


def update_group_report(path, filename, usuario):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [norm_key(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    def idx(name_options):
        for opt in name_options:
            nk = norm_key(opt)
            if nk in headers:
                return headers.index(nk)
        return None

    i_nf = idx(["NF", "Nota Fiscal"])
    i_status = idx(["Status Atualizacao", "Status", "Status Grupo"])
    i_occ = idx(["Ocorrencia Atualizacao", "Ocorrencia", "Ocorrência"])
    i_resp = idx(["Responsavel", "Responsável", "Usuario", "Usuário"])

    if i_nf is None or i_status is None:
        raise ValueError("A planilha de retorno precisa ter as colunas NF e Status Atualizacao.")

    conn = connect()
    try:
        upload_id = save_upload_record(conn, "RETORNO_GRUPO", filename, path, usuario)
        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nf = norm_text(row[i_nf] if i_nf < len(row) else "")
            status = norm_text(row[i_status] if i_status < len(row) else "")
            occ = norm_text(row[i_occ] if i_occ is not None and i_occ < len(row) else "")
            resp = norm_text(row[i_resp] if i_resp is not None and i_resp < len(row) else usuario)
            if not nf or not status:
                continue

            status_key = norm_key(status)
            finalizado = 1 if "ENTREGUE" in status_key or "CANCEL" in status_key else 0
            cancelado = 1 if "CANCEL" in status_key else 0

            # Se usuario for rastreamento, limita ao proprio grupo.
            u = conn.execute("SELECT * FROM users WHERE login=?", (usuario,)).fetchone()
            params = [status, occ, resp or usuario, usuario, datetime.now().isoformat(), finalizado, cancelado, nf]
            sql = """
                UPDATE demandas
                SET status_grupo=?, ocorrencia_grupo=?, responsavel_login=?, updated_by=?, updated_at=?,
                    finalizado=CASE WHEN ?=1 THEN 1 ELSE finalizado END,
                    cancelado=CASE WHEN ?=1 THEN 1 ELSE cancelado END
                WHERE active=1 AND nf=?
            """
            if u and u["perfil"] == "rastreamento" and u["grupo_id"]:
                sql += " AND grupo_id=?"
                params.append(u["grupo_id"])
            cur = conn.execute(sql, params)
            updated += cur.rowcount

        conn.commit()
        return updated, upload_id
    finally:
        conn.close()


def demand_where_for_user(user, alias="d"):
    perfil = user.get("perfil")
    if perfil in ("admin", "gestor", "rastreamento"):
        if perfil == "rastreamento":
            return f"{alias}.active=1 AND {alias}.grupo_id={int(user.get('grupo_id') or 0)}"
        return f"{alias}.active=1"
    if perfil == "cliente":
        cod = user.get("cliente_codigo") or ""
        safe = cod.replace("'", "''")
        return f"{alias}.active=1 AND {alias}.cliente_codigo='{safe}'"
    if perfil == "expedicao":
        return f"{alias}.active=1"
    return f"{alias}.active=1 AND 1=0"


def fetch_demands(user, filters=None, limit=MAX_QUERY_ROWS):
    filters = filters or {}
    where = [demand_where_for_user(user, "d")]
    params = []

    for key in ("cliente_codigo", "grupo_id", "parceiro", "estado"):
        val = norm_text(filters.get(key, ""))
        if val:
            if key == "grupo_id":
                where.append("d.grupo_id=?")
                params.append(int(val))
            else:
                where.append(f"d.{key}=?")
                params.append(val)

    ano = to_int(filters.get("ano"), 0)
    mes = to_int(filters.get("mes"), 0)
    if ano:
        where.append("d.ano=?")
        params.append(ano)
    if mes:
        where.append("d.mes=?")
        params.append(mes)

    if filters.get("nf"):
        where.append("d.nf LIKE ?")
        params.append("%" + norm_text(filters["nf"]) + "%")

    sql = f"""
        SELECT d.*, g.nome AS grupo_nome
        FROM demandas d
        LEFT JOIN grupos g ON g.id=d.grupo_id
        WHERE {' AND '.join(where)}
        ORDER BY d.dias_atraso_sla DESC, d.data_prevista ASC, d.id DESC
        LIMIT ?
    """
    params.append(limit)
    conn = connect()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def make_summary(rows):
    total = len(rows)
    em_aberto = [r for r in rows if not r.get("finalizado")]
    prazo = [r for r in rows if r.get("status_sistema") == "Em aberto no prazo"]
    atraso = [r for r in rows if r.get("status_sistema") in ("Em aberto com atraso", "Entregue atrasado")]
    cancelados = [r for r in rows if r.get("cancelado")]
    ocorr = [r for r in rows if r.get("status_sistema") == "Em aberto com ocorrência" or norm_text(r.get("ocorrencia_grupo"))]
    ok_sla = [r for r in rows if r.get("status_sistema") in ("Entregue no prazo", "Em aberto no prazo")]
    return {
        "total": total,
        "em_aberto": len(em_aberto),
        "prazo": len(prazo),
        "atraso": len(atraso),
        "cancelados": len(cancelados),
        "ocorrencias": len(ocorr),
        "sla": pct(len(ok_sla), total),
        "pct_em_aberto": pct(len(em_aberto), total),
        "pct_prazo": pct(len(prazo), total),
        "pct_atraso": pct(len(atraso), total),
        "pct_cancelados": pct(len(cancelados), total),
        "pct_ocorrencias": pct(len(ocorr), total),
    }


def grouped_summary(rows, key):
    groups = {}
    for r in rows:
        k = norm_text(r.get(key)) or "Nao informado"
        groups.setdefault(k, []).append(r)
    out = []
    for k, itens in groups.items():
        s = make_summary(itens)
        out.append((k, s))
    out.sort(key=lambda x: (x[1]["pct_atraso"], x[1]["atraso"], x[1]["total"]), reverse=True)
    return out


def select_options(table, value_col, label_col=None, selected=""):
    conn = connect()
    label_col = label_col or value_col
    rows = conn.execute(f"SELECT {value_col} v, {label_col} l FROM {table} WHERE ativo=1 ORDER BY {label_col}").fetchall()
    conn.close()
    htmls = ["<option value=''>Selecione</option>"]
    for r in rows:
        sel = " selected" if str(r["v"]) == str(selected) else ""
        htmls.append(f"<option value='{html_escape(r['v'])}'{sel}>{html_escape(r['l'])}</option>")
    return "".join(htmls)


def distinct_options(field, user, selected="", label_all="Todos"):
    rows = fetch_demands(user, {}, limit=50000)
    vals = sorted(set(norm_text(r.get(field)) for r in rows if norm_text(r.get(field))))
    out = [f"<option value=''>{html_escape(label_all)}</option>"]
    for v in vals:
        sel = " selected" if v == selected else ""
        out.append(f"<option value='{html_escape(v)}'{sel}>{html_escape(v)}</option>")
    return "".join(out)


def distinct_cliente_options(user, selected=""):
    where = demand_where_for_user(user, "d")
    conn = connect()
    rows = conn.execute(f"""
        SELECT cliente_codigo, MAX(cliente_nome) AS cliente_nome, COUNT(*) AS qtd
        FROM demandas d
        WHERE {where} AND cliente_codigo<>''
        GROUP BY cliente_codigo
        ORDER BY cliente_nome, cliente_codigo
    """).fetchall()
    conn.close()
    out = ["<option value=''>Todos clientes</option>"]
    for r in rows:
        cod = norm_text(r["cliente_codigo"])
        nome = norm_text(r["cliente_nome"])
        label = f"{cod} - {nome} ({r['qtd']})" if nome else f"{cod} ({r['qtd']})"
        sel = " selected" if cod == selected else ""
        out.append(f"<option value='{html_escape(cod)}'{sel}>{html_escape(label)}</option>")
    return "".join(out)




def current_year_month():
    hoje = date.today()
    return hoje.year, hoje.month


def month_year_from_qs(qs):
    qs = qs or {}
    raw_ano = qs.get("ano", [""])
    raw_mes = qs.get("mes", [""])
    y = to_int(raw_ano[0] if isinstance(raw_ano, list) else raw_ano, 0)
    m = to_int(raw_mes[0] if isinstance(raw_mes, list) else raw_mes, 0)
    if y <= 0 or m <= 0:
        y, m = current_year_month()
    if m < 1 or m > 12:
        m = date.today().month
    return y, m


def month_filter_html(action, qs=None):
    ano_atual, mes_atual = month_year_from_qs(qs or {})
    nomes_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    anos = set([date.today().year])
    try:
        conn = connect()
        for r in conn.execute("SELECT DISTINCT ano FROM demandas WHERE ano IS NOT NULL ORDER BY ano DESC").fetchall():
            if r["ano"]:
                anos.add(int(r["ano"]))
        conn.close()
    except Exception:
        pass
    opt_anos = "".join(f"<option value='{a}'{' selected' if a == ano_atual else ''}>{a}</option>" for a in sorted(anos, reverse=True))
    opt_meses = "".join(
        f"<option value='{i}'{' selected' if i == mes_atual else ''}>{i:02d} - {nomes_meses[i-1]}</option>"
        for i in range(1, 13)
    )
    return f"""
    <section class='card month_filter_card'>
        <form method='get' action='{html_escape(action)}' class='month_filter_form'>
            <div><label>Mês de controle</label><select name='mes'>{opt_meses}</select></div>
            <div><label>Ano</label><select name='ano'>{opt_anos}</select></div>
            <button>Aplicar mês</button>
        </form>
    </section>
    """


def month_options_html(selected_mes=None):
    selected_mes = to_int(selected_mes, date.today().month)
    nomes_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    return "".join(
        f"<option value='{i}'{' selected' if i == selected_mes else ''}>{i:02d} - {nomes_meses[i-1]}</option>"
        for i in range(1, 13)
    )


def kpi_card(title, qtd, perc, cls):
    return f"""
    <div class='kpi {cls}'>
        <b>{qtd}</b>
        <span>{html_escape(title)}</span>
        <small>{perc:.2f}%</small>
    </div>
    """


def table_demands(rows, show_group=True, limit=300):
    trs = []
    for r in rows[:limit]:
        trs.append(f"""
        <tr>
            <td>{html_escape(r.get('nf'))}</td>
            <td>{html_escape(r.get('status_grupo') or r.get('status_sistema'))}</td>
            <td>{html_escape(r.get('grupo_nome')) if show_group else ''}</td>
            <td>{html_escape(r.get('parceiro'))}</td>
            <td>{html_escape(r.get('estado'))}</td>
            <td>{html_escape(r.get('cliente_nome'))}</td>
            <td>{fmt_date(r.get('data_prevista'))}</td>
            <td>{r.get('dias_atraso_sla') or 0}</td>
            <td>{r.get('atraso_atualizacao_dias') or 0}</td>
            <td>{html_escape(r.get('responsavel_login'))}</td>
            <td>{html_escape(r.get('ocorrencia_grupo') or r.get('ocorrencia_original'))}</td>
        </tr>
        """)
    return f"""
    <div class='tablebox'>
        <table>
            <thead><tr>
                <th>NF</th><th>Status</th><th>Grupo</th><th>Parceiro</th><th>UF</th><th>Cliente</th>
                <th>Data Prevista</th><th>Atraso SLA</th><th>Atraso Atualiz.</th><th>Responsavel</th><th>Ocorrencia</th>
            </tr></thead>
            <tbody>{''.join(trs)}</tbody>
        </table>
    </div>
    """


def bar_html(label, value, total):
    p = pct(value, total)
    return f"""
    <div class='barline'>
        <div><b>{html_escape(label)}</b><span>{value} | {p:.2f}%</span></div>
        <div class='bar'><i style='width:{min(100,p)}%'></i></div>
    </div>
    """


def nav(user):
    menu = [
        ("admin", "/", "Admin"),
        ("acompanhamento", "/acompanhamento", "Acompanhamento"),
        ("rastreamento", "/rastreamento", "Rastreamento"),
        ("chamado_parceiro", "/chamado_parceiro", "Chamado Parceiro"),
        ("cliente", "/cliente", "Cliente"),
        ("expedicao", "/expedicao", "Expedição"),
    ]
    items = [(href, label) for perm, href, label in menu if has_permission(user, (perm,))]
    links = "".join(f"<a href='{href}'>{label}</a>" for href, label in items)
    return f"<nav class='nav'>{links}<a href='/alterar_senha'>Senha</a><a href='/logout'>Sair</a></nav>"

def page(title, body, user=None):
    userbar = ""
    if user:
        userbar = f"""
        <div class='top'>
            <div><h1>{html_escape(title)}</h1><p>Usuário: <b>{html_escape(user.get('nome'))}</b> | Perfil: <b>{html_escape(user.get('perfil'))}</b></p></div>
            {nav(user)}
        </div>
        """
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>{html_escape(title)}</title>{CSS}</head><body><main class='wrap wide'>{userbar}{body}</main></body></html>"""


def render_login(msg=""):
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Login Rastreamento</title>{CSS}</head><body>
    <div class='login_wrap'>
        <div class='login_card'>
            <div class='brand'><h1>Rastreamento PRO</h1><p>Gestão por grupo, parceiro e SLA</p></div>
            {'<div class="err">'+html_escape(msg)+'</div>' if msg else ''}
            <form method='post' action='/login' class='form'>
                <label>Usuário</label><input name='login' required autocomplete='username'>
                <label>Senha</label><input type='password' name='senha' required autocomplete='current-password'>
                <button>Entrar</button>
            </form>
            <p class='note'>Login inicial: admin / Admin@123</p>
        </div>
    </div></body></html>"""


def render_change_password(user, msg=""):
    body = f"""
    <section class='card narrow'>
        {'<div class="err">'+html_escape(msg)+'</div>' if msg else ''}
        <form method='post' action='/alterar_senha' class='form'>
            <label>Senha atual</label><input type='password' name='senha_atual' required>
            <label>Nova senha</label><input type='password' name='nova_senha' minlength='6' required>
            <label>Confirmar nova senha</label><input type='password' name='confirmar_senha' minlength='6' required>
            <button>Salvar senha</button>
        </form>
    </section>
    """
    return page("Alterar senha", body, user)


def render_admin(user, msg=""):
    body = f"""
    {'<div class="okmsg">'+html_escape(msg)+'</div>' if msg else ''}
    <section class='grid two'>
        <div class='card'>
            <h2>Upload relatório geral</h2>
            <p>Responsabilidade do Admin/Gestor. O novo upload substitui a base ativa, mas mantém o histórico anual no banco.</p>
            <form method='post' action='/upload_geral' enctype='multipart/form-data' class='upload'>
                <div class='month_filter_form upload_month_form'>
                    <div>
                        <label>Mês da planilha</label>
                        <select name='mes_controle'>{month_options_html()}</select>
                    </div>
                    <div>
                        <label>Ano da planilha</label>
                        <input name='ano_controle' value='{date.today().year}' inputmode='numeric'>
                    </div>
                </div>
                <input type='file' name='file' accept='.xlsx,.xlsm' required>
                <button>Anexar relatório geral</button>
            </form>
        </div>
        <div class='card'>
            <h2>Regras desta versão</h2>
            <p>Não deixa a base ativa anterior misturada com a nova. O passado fica salvo para indicadores de Janeiro a Dezembro.</p>
            <p>Parceiro tende a ficar no mesmo grupo. Se não houver grupo, entra no grupo com menor demanda.</p>
        </div>
    </section>
    <section class='grid four'>
        <a class='module' href='/cliente'>Cliente<br><small>visualização filtrada</small></a>
        <a class='module' href='/'>Admin<br><small>controle total</small></a>
        <a class='module active' href='/rastreamento'>Rastreamento<br><small>dashboard operacional</small></a>
        <a class='module' href='/chamado_parceiro'>Chamado Parceiro<br><small>SLA e relatórios</small></a>
        <a class='module' href='/expedicao'>Expedição<br><small>base criada</small></a>
    </section>
    <section class='grid two'>
        <a class='module admin_tool' href='/usuarios'>Usuários<br><small>criar login e vincular perfil</small></a>
        <a class='module admin_tool' href='/grupos'>Grupos<br><small>criar grupos e responsáveis</small></a>
    </section>
    {render_perfis_admin()}
    """
    return page("Admin - Controle Total", body, user)


def render_gestor(user, qs=None):
    ano, mes = month_year_from_qs(qs or {})
    rows = fetch_demands(user, {"ano": ano, "mes": mes})
    s = make_summary(rows)
    body = f"""
    {month_filter_html('/gestor', qs)}
    <section class='grid kpis'>
        {kpi_card('Total ativo', s['total'], 100 if s['total'] else 0, 'blue')}
        {kpi_card('SLA geral', f"{s['sla']:.2f}%", s['sla'], 'green' if s['sla'] >= 95 else 'red')}
        {kpi_card('Em aberto', s['em_aberto'], s['pct_em_aberto'], 'orange')}
        {kpi_card('Atrasados', s['atraso'], s['pct_atraso'], 'red')}
        {kpi_card('Ocorrências', s['ocorrencias'], s['pct_ocorrencias'], 'purple')}
    </section>
    <section class='card'>
        <h2>Upload relatório geral</h2>
        <form method='post' action='/upload_geral' enctype='multipart/form-data' class='upload'>
            <div class='month_filter_form upload_month_form'>
                <div>
                    <label>Mês da planilha</label>
                    <select name='mes_controle'>{month_options_html(mes)}</select>
                </div>
                <div>
                    <label>Ano da planilha</label>
                    <input name='ano_controle' value='{ano}' inputmode='numeric'>
                </div>
            </div>
            <input type='file' name='file' accept='.xlsx,.xlsm' required>
            <button>Anexar relatório geral</button>
        </form>
    </section>
    """
    return page("Gestor - Visão Geral", body, user)




def acompanhamento_summary(rows):
    total_base = len(rows)
    em_aberto = [r for r in rows if not r.get("finalizado")]
    aberto_no_prazo = [r for r in em_aberto if norm_text(r.get("status_sistema")) == "Em aberto no prazo"]
    aberto_fora_prazo = [
        r for r in em_aberto
        if norm_text(r.get("status_sistema")) in ("Em aberto com atraso", "Em aberto com ocorrência")
        or (r.get("dias_atraso_sla") or 0) > 0
    ]
    finalizado_no_prazo = [r for r in rows if norm_text(r.get("status_sistema")) == "Entregue no prazo"]
    finalizado_atrasado = [r for r in rows if norm_text(r.get("status_sistema")) == "Entregue atrasado"]
    ocorrencias_aberto = [
        r for r in em_aberto
        if norm_text(r.get("status_sistema")) == "Em aberto com ocorrência"
        or norm_text(r.get("ocorrencia_grupo"))
        or norm_text(r.get("ocorrencia_original"))
    ]
    return {
        "total_base": total_base,
        "em_aberto": len(em_aberto),
        "aberto_no_prazo": len(aberto_no_prazo),
        "aberto_fora_prazo": len(aberto_fora_prazo),
        "finalizado_no_prazo": len(finalizado_no_prazo),
        "finalizado_atrasado": len(finalizado_atrasado),
        "ocorrencias_aberto": len(ocorrencias_aberto),
        "pct_em_aberto": pct(len(em_aberto), total_base),
        "pct_aberto_no_prazo": pct(len(aberto_no_prazo), total_base),
        "pct_aberto_fora_prazo": pct(len(aberto_fora_prazo), total_base),
        "pct_finalizado_no_prazo": pct(len(finalizado_no_prazo), total_base),
        "pct_finalizado_atrasado": pct(len(finalizado_atrasado), total_base),
        "pct_ocorrencias_aberto": pct(len(ocorrencias_aberto), total_base),
    }


def render_acompanhamento(user, qs=None):
    ano, mes = month_year_from_qs(qs or {})
    rows = fetch_demands(user, {"ano": ano, "mes": mes}, limit=MAX_QUERY_ROWS)
    s = acompanhamento_summary(rows)
    rows_operacionais = [r for r in rows if not r.get("finalizado")]
    parceiros = grouped_summary(rows_operacionais, "parceiro")
    grupos = grouped_summary(rows_operacionais, "grupo_nome")
    usuarios = grouped_summary(rows_operacionais, "responsavel_login")

    body = f"""
    {month_filter_html('/acompanhamento', qs)}
    <section class='card'>
        <h2>Tela de acompanhamento</h2>
        <p class='note'>Visualização geral da base ativa, sem filtros. Indicadores separados entre demandas em aberto e finalizadas.</p>
    </section>

    <section class='grid six acompanhamento_grid'>
        {kpi_card('Total de pedidos em aberto', s['em_aberto'], s['pct_em_aberto'], 'blue')}
        {kpi_card('Pedidos em aberto - No prazo', s['aberto_no_prazo'], s['pct_aberto_no_prazo'], 'green')}
        {kpi_card('Pedidos em aberto - Fora do prazo', s['aberto_fora_prazo'], s['pct_aberto_fora_prazo'], 'red')}
        {kpi_card('Finalizado - No prazo', s['finalizado_no_prazo'], s['pct_finalizado_no_prazo'], 'green')}
        {kpi_card('Finalizado - Entregue atrasado', s['finalizado_atrasado'], s['pct_finalizado_atrasado'], 'orange')}
        {kpi_card('Ocorrências em aberto', s['ocorrencias_aberto'], s['pct_ocorrencias_aberto'], 'purple')}
    </section>

    <section class='card'>
        <h2>Resumo da base ativa</h2>
        <div class='tablebox small'>
            <table>
                <tr><th>Indicador</th><th>Quantidade</th><th>% sobre base ativa</th></tr>
                <tr><td>Total da base ativa</td><td>{s['total_base']}</td><td>100.00%</td></tr>
                <tr><td>Total de pedidos em aberto</td><td>{s['em_aberto']}</td><td>{s['pct_em_aberto']:.2f}%</td></tr>
                <tr><td>Pedidos em aberto - No prazo</td><td>{s['aberto_no_prazo']}</td><td>{s['pct_aberto_no_prazo']:.2f}%</td></tr>
                <tr><td>Pedidos em aberto - Fora do prazo</td><td>{s['aberto_fora_prazo']}</td><td>{s['pct_aberto_fora_prazo']:.2f}%</td></tr>
                <tr><td>Finalizado - No prazo</td><td>{s['finalizado_no_prazo']}</td><td>{s['pct_finalizado_no_prazo']:.2f}%</td></tr>
                <tr><td>Finalizado - Entregue atrasado</td><td>{s['finalizado_atrasado']}</td><td>{s['pct_finalizado_atrasado']:.2f}%</td></tr>
                <tr><td>Ocorrências em aberto</td><td>{s['ocorrencias_aberto']}</td><td>{s['pct_ocorrencias_aberto']:.2f}%</td></tr>
            </table>
        </div>
    </section>

    <section class='card'>
        <h2>Parceiros críticos - em aberto</h2>
        {table_group_small(parceiros)}
    </section>

    <section class='grid two'>
        <div class='card'>
            <h2>Performance dos grupos - em aberto</h2>
            {table_group_small(grupos)}
        </div>
        <div class='card'>
            <h2>Performance dos usuários - em aberto</h2>
            {table_group_small(usuarios)}
        </div>
    </section>
    """
    return page("Tela de acompanhamento", body, user)


def render_rastreamento(user, qs=None):
    qs = qs or {}
    ano, mes = month_year_from_qs(qs or {})
    filters = {
        "cliente_codigo": qs.get("cliente", [""])[0],
        "grupo_id": qs.get("grupo", [""])[0],
        "parceiro": qs.get("parceiro", [""])[0],
        "estado": qs.get("estado", [""])[0],
        "ano": ano,
        "mes": mes,
    }
    rows = fetch_demands(user, filters)
    # Base total filtrada fica nos KPIs gerais; distribuição/performance trabalha só com demandas em aberto.
    rows_operacionais = [r for r in rows if not r.get("finalizado")]
    s = make_summary(rows)
    grupo_nome = ""
    if user.get("perfil") == "rastreamento" and user.get("grupo_id"):
        conn = connect()
        g = conn.execute("SELECT nome FROM grupos WHERE id=?", (user.get("grupo_id"),)).fetchone()
        grupo_nome = g["nome"] if g else ""
        conn.close()

    grupos = grouped_summary(rows_operacionais, "grupo_nome")
    parceiros = grouped_summary(rows_operacionais, "parceiro")
    estados = grouped_summary(rows_operacionais, "estado")

    grupos_rows = "".join(f"""
        <tr><td>{html_escape(k)}</td><td>{v['total']}</td><td>{v['prazo']} ({v['pct_prazo']:.2f}%)</td><td>{v['atraso']} ({v['pct_atraso']:.2f}%)</td><td>{v['cancelados']} ({v['pct_cancelados']:.2f}%)</td><td>{v['ocorrencias']} ({v['pct_ocorrencias']:.2f}%)</td><td>{v['sla']:.2f}%</td></tr>
    """ for k, v in grupos[:50])

    filtros_html = ""
    if user.get("perfil") in ("admin", "gestor"):
        filtros_html = f"""
        <section class='card filtros_sla_card'>
            <form method='get' action='/rastreamento' class='filter_form filter_form_sla'>
                <div class='filtros_sla_title'><h2>Filtros de SLA</h2></div>
                <div class='campo_mes'><label>Mês</label><select name='mes'>{month_options_html(mes)}</select></div>
                <div class='campo_ano'><label>Ano</label><input name='ano' value='{ano}'></div>
                <div class='campo_cliente'><label>Cliente (AF/AG)</label><select name='cliente'>{distinct_cliente_options(user, filters['cliente_codigo'])}</select></div>
                <div class='campo_grupo'><label>Grupo</label><select name='grupo'><option value=''>Todos grupos</option>{select_options('grupos','id','nome',filters['grupo_id'])}</select></div>
                <div class='campo_parceiro'><label>Parceiro</label><select name='parceiro'>{distinct_options('parceiro', user, filters['parceiro'], 'Todos parceiros')}</select></div>
                <div class='campo_estado'><label>UF/Estado</label><select name='estado'>{distinct_options('estado', user, filters['estado'], 'Todos estados')}</select></div>
                <div class='filter_actions'><button>Aplicar</button><a class='btn secondary' href='/rastreamento'>Limpar</a></div>
            </form>
        </section>
        """

    download_links = f"""
    <section class='card'>
        <h2>Download e retorno do grupo {html_escape(grupo_nome)}</h2>
        <div class='actions left'>
            <a class='btn' href='/download_grupo?modo=meu'>Baixar minha demanda</a>
            <a class='btn secondary' href='/download_grupo?modo=grupo'>Baixar demanda do grupo</a>
        </div>
        <form method='post' action='/upload_retorno' enctype='multipart/form-data' class='upload'>
            <input type='file' name='file' accept='.xlsx,.xlsm' required>
            <button>Subir planilha atualizada</button>
        </form>
        <p class='note'>A planilha baixada já vem com colunas de Status Atualizacao e Ocorrencia Atualizacao.</p>
    </section>
    """

    admin_download = ""
    if user.get("perfil") in ("admin", "gestor"):
        admin_download = """
        <section class='card'>
            <h2>Redistribuição operacional</h2>
            <p>A distribuição operacional considera somente demandas em aberto. Entregues, canceladas e finalizadas ficam fora da fila dos grupos.</p>
            <form method='post' action='/redistribuir_demanda' class='actions left'><button>Gerar/Redistribuir demanda em aberto</button><a class='btn secondary' href='/download_todos'>Baixar base ativa completa</a></form>
        </section>
        """

    body = f"""
    {month_filter_html('/rastreamento', qs) if user.get('perfil') == 'rastreamento' else ''}
    {filtros_html}
    <section class='grid kpis'>
        {kpi_card('Total de pedidos em aberto', s['em_aberto'], s['pct_em_aberto'], 'blue')}
        {kpi_card('Pedidos em aberto no prazo', s['prazo'], s['pct_prazo'], 'green')}
        {kpi_card('Pedidos em atraso', s['atraso'], s['pct_atraso'], 'red')}
        {kpi_card('Pedidos cancelados', s['cancelados'], s['pct_cancelados'], 'gray')}
        {kpi_card('Em aberto com ocorrência', s['ocorrencias'], s['pct_ocorrencias'], 'purple')}
    </section>

    <section class='grid two'>
        <div class='card'>
            <h2>Evolução por grupo - somente demandas em aberto</h2>
            <div class='tablebox small'>
                <table><tr><th>Grupo</th><th>Total</th><th>No prazo</th><th>Atraso</th><th>Cancelado</th><th>Ocorr.</th><th>SLA</th></tr>{grupos_rows}</table>
            </div>
        </div>
        <div class='card'>
            <h2>SLA geral da empresa</h2>
            {bar_html('SLA geral', int(s['sla']), 100)}
            {bar_html('Atrasados', s['atraso'], max(1,s['total']))}
            {bar_html('Ocorrências', s['ocorrencias'], max(1,s['total']))}
            {bar_html('Cancelados', s['cancelados'], max(1,s['total']))}
        </div>
    </section>

    <section class='grid two'>
        <div class='card'><h2>Parceiros críticos - em aberto</h2>{table_group_small(parceiros)}</div>
        <div class='card'><h2>SLA por estado - em aberto</h2>{table_group_small(estados)}</div>
    </section>

    {download_links if user.get('perfil') == 'rastreamento' else ''}
    {admin_download}

    <section class='card'>
        <h2>Demandas em aberto detalhadas</h2>
        <p class='note'>Esta lista operacional mostra apenas o que ainda precisa de ação. Fechados ficam no banco/histórico e nos indicadores gerais.</p>
        {table_demands(rows_operacionais)}
    </section>
    """
    return page("Rastreamento - Dashboard Operacional", body, user)


def table_group_small(grouped):
    trs = []
    for k, s in grouped[:20]:
        cls = "bad" if s["pct_atraso"] >= 20 else "warn" if s["pct_atraso"] >= 10 else "ok"
        trs.append(f"<tr><td>{html_escape(k)}</td><td>{s['total']}</td><td>{s['atraso']}</td><td><span class='badge {cls}'>{s['pct_atraso']:.2f}%</span></td><td>{s['sla']:.2f}%</td></tr>")
    return f"<div class='tablebox small'><table><tr><th>Nome</th><th>Total</th><th>Atraso</th><th>% Atraso</th><th>SLA</th></tr>{''.join(trs)}</table></div>"



def query_string_from_filters(filters):
    data = {}
    for k, v in (filters or {}).items():
        if v is not None and str(v).strip() != "":
            data[k] = str(v).strip()
    return urlencode(data)


def parceiro_filters_from_qs(qs):
    ano, mes = month_year_from_qs(qs or {})
    return {
        "ano": ano,
        "mes": mes,
        "cliente_codigo": (qs or {}).get("cliente", [""])[0],
        "grupo_id": (qs or {}).get("grupo", [""])[0],
        "parceiro": (qs or {}).get("parceiro", [""])[0],
        "estado": (qs or {}).get("estado", [""])[0],
    }


def parceiro_report_rows(user, filters, tipo):
    base_filters = dict(filters or {})
    rows = fetch_demands(user, base_filters, limit=MAX_QUERY_ROWS)
    tipo = norm_text(tipo)
    ontem = date.today() - timedelta(days=1)

    if tipo == "1":
        # Relatório 1: tudo com data de emissão CTE/Minuta do dia anterior,
        # mantendo a visão de conformidade SLA pelo status calculado no sistema.
        rows = [r for r in rows if to_date(r.get("data_inicial")) == ontem]
    elif tipo == "2":
        rows = [r for r in rows if not r.get("finalizado")]
    elif tipo == "3":
        rows = [r for r in rows if not r.get("finalizado") and norm_text(r.get("status_sistema")) == "Em aberto no prazo"]
    elif tipo == "4":
        rows = [
            r for r in rows
            if not r.get("finalizado") and (
                norm_text(r.get("status_sistema")) in ("Em aberto com atraso", "Em aberto com ocorrência")
                or (r.get("dias_atraso_sla") or 0) > 0
            )
        ]
    elif tipo == "5":
        # Relatório geral do mês: mantém todos os status da demanda do mês filtrado.
        rows = rows
    else:
        rows = []
    return rows


def parceiro_report_title(tipo):
    nomes = {
        "1": "Conforme SLA e emissão CTE/Minuta - dia anterior",
        "2": "Pedidos em aberto por parceiro",
        "3": "Pedidos em aberto no prazo por parceiro",
        "4": "Pedidos em aberto SLA atrasado por parceiro",
        "5": "Geral mensal com status por parceiro",
    }
    return nomes.get(str(tipo), "Relatório Parceiro")


def create_partner_report_workbook(rows, title="Relatorio Parceiro"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    headers = [
        "Parceiro", "NF", "Status Sistema", "Status Grupo", "Cliente Código", "Cliente Nome",
        "UF/Estado", "Cidade Destino", "Data Emissão CTE/Minuta", "Data Prevista", "Data Entrega",
        "SLA", "Dias Atraso SLA", "Grupo", "Responsável", "Ocorrência Original", "Ocorrência Grupo",
        "Filial", "Tipo", "Link Comprovante"
    ]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("parceiro"), r.get("nf"), r.get("status_sistema"), r.get("status_grupo"),
            r.get("cliente_codigo"), r.get("cliente_nome"), r.get("estado"), r.get("cidade_destino"),
            fmt_date(r.get("data_inicial")), fmt_date(r.get("data_prevista")), fmt_date(r.get("data_entrega")),
            r.get("sla") or 0, r.get("dias_atraso_sla") or 0, r.get("grupo_nome"),
            r.get("responsavel_login"), r.get("ocorrencia_original"), r.get("ocorrencia_grupo"),
            r.get("filial"), r.get("tipo"), r.get("link_comprovante")
        ])

    fill_title = PatternFill("solid", fgColor="0B6B8F")
    fill_header = PatternFill("solid", fgColor="E7F3EC")
    font_title = Font(color="FFFFFF", bold=True, size=14)
    font_header = Font(color="102B1D", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    ws[1][0].fill = fill_title
    ws[1][0].font = font_title
    ws[1][0].alignment = Alignment(horizontal="center")
    for cell_obj in ws[2]:
        cell_obj.fill = fill_header
        cell_obj.font = font_header
        cell_obj.alignment = Alignment(horizontal="center")
        cell_obj.border = Border(bottom=thin)
    widths = [30,18,24,24,16,42,12,28,22,16,16,10,16,20,18,36,36,22,18,42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def render_chamado_parceiro(user, qs=None):
    qs = qs or {}
    filters = parceiro_filters_from_qs(qs)
    rows = fetch_demands(user, filters, limit=MAX_QUERY_ROWS)
    rows_operacionais = [r for r in rows if not r.get("finalizado")]
    s = make_summary(rows)
    parceiros = grouped_summary(rows_operacionais, "parceiro")
    base_qs = query_string_from_filters({
        "ano": filters.get("ano"),
        "mes": filters.get("mes"),
        "cliente": filters.get("cliente_codigo"),
        "grupo": filters.get("grupo_id"),
        "parceiro": filters.get("parceiro"),
        "estado": filters.get("estado"),
    })
    def report_link(tipo, label):
        sep = "&" if base_qs else ""
        return f"<a class='btn report_btn' href='/download_relatorio_parceiro?tipo={tipo}{sep}{base_qs}'>{html_escape(label)}</a>"

    body = f"""
    <section class='card filtros_sla_card'>
        <form method='get' action='/chamado_parceiro' class='filter_form filter_form_sla'>
            <div class='filtros_sla_title'><h2>Chamado Parceiro</h2></div>
            <div class='campo_mes'><label>Mês</label><select name='mes'>{month_options_html(filters['mes'])}</select></div>
            <div class='campo_ano'><label>Ano</label><input name='ano' value='{filters['ano']}'></div>
            <div class='campo_cliente'><label>Cliente (AF/AG)</label><select name='cliente'>{distinct_cliente_options(user, filters['cliente_codigo'])}</select></div>
            <div class='campo_grupo'><label>Grupo</label><select name='grupo'><option value=''>Todos grupos</option>{select_options('grupos','id','nome',filters['grupo_id'])}</select></div>
            <div class='campo_parceiro'><label>Parceiro</label><select name='parceiro'>{distinct_options('parceiro', user, filters['parceiro'], 'Todos parceiros')}</select></div>
            <div class='campo_estado'><label>UF/Estado</label><select name='estado'>{distinct_options('estado', user, filters['estado'], 'Todos estados')}</select></div>
            <div class='filter_actions'><button>Aplicar</button><a class='btn secondary' href='/chamado_parceiro'>Limpar</a></div>
        </form>
    </section>

    <section class='grid kpis'>
        {kpi_card('Total no filtro', s['total'], 100 if s['total'] else 0, 'blue')}
        {kpi_card('SLA geral', f"{s['sla']:.2f}%", s['sla'], 'green' if s['sla'] >= 95 else 'red')}
        {kpi_card('Pedidos em aberto', s['em_aberto'], s['pct_em_aberto'], 'orange')}
        {kpi_card('Em aberto no prazo', s['prazo'], s['pct_prazo'], 'green')}
        {kpi_card('SLA atrasado', s['atraso'], s['pct_atraso'], 'red')}
    </section>

    <section class='card'>
        <h2>Baixar relatórios por parceiro</h2>
        <div class='actions left report_actions'>
            {report_link('1', '1 - Conforme SLA + emissão dia anterior')}
            {report_link('2', '2 - Todos pedidos em aberto')}
            {report_link('3', '3 - Em aberto no prazo')}
            {report_link('4', '4 - Em aberto SLA atrasado')}
            {report_link('5', '5 - Geral mensal por parceiro')}
        </div>
        <p class='note'>Os relatórios respeitam os filtros acima: mês, ano, cliente, grupo, parceiro e UF/Estado.</p>
    </section>

    <section class='card'>
        <h2>Parceiros críticos - em aberto</h2>
        {table_group_small(parceiros)}
    </section>

    <section class='card'>
        <h2>Consulta SLA/Performance por parceiro</h2>
        {table_demands(rows_operacionais)}
    </section>
    """
    return page("Chamado Parceiro", body, user)

def render_cliente(user, qs=None):
    ano, mes = month_year_from_qs(qs or {})
    rows = fetch_demands(user, {"ano": ano, "mes": mes})
    s = make_summary(rows)
    body = f"""
    {month_filter_html('/cliente', qs)}
    <section class='grid kpis'>
        {kpi_card('Total do cliente', s['total'], 100 if s['total'] else 0, 'blue')}
        {kpi_card('SLA', f"{s['sla']:.2f}%", s['sla'], 'green' if s['sla'] >= 95 else 'red')}
        {kpi_card('Em aberto', s['em_aberto'], s['pct_em_aberto'], 'orange')}
        {kpi_card('Atrasados', s['atraso'], s['pct_atraso'], 'red')}
        {kpi_card('Ocorrências', s['ocorrencias'], s['pct_ocorrencias'], 'purple')}
    </section>
    <section class='card'><h2>Demandas do cliente</h2>{table_demands(rows)}</section>
    """
    return page("Cliente - Visualização Restrita", body, user)


def render_expedicao(user, qs=None):
    ano, mes = month_year_from_qs(qs or {})
    rows = fetch_demands(user, {"ano": ano, "mes": mes})
    s = make_summary(rows)
    body = f"""
    {month_filter_html('/expedicao', qs)}
    <section class='grid kpis'>
        {kpi_card('Total ativo no mês', s['total'], 100 if s['total'] else 0, 'blue')}
        {kpi_card('Em aberto', s['em_aberto'], s['pct_em_aberto'], 'orange')}
        {kpi_card('No prazo', s['prazo'], s['pct_prazo'], 'green')}
        {kpi_card('Atrasados', s['atraso'], s['pct_atraso'], 'red')}
        {kpi_card('Ocorrências', s['ocorrencias'], s['pct_ocorrencias'], 'purple')}
    </section>
    <section class='card'>
        <h2>Expedição</h2>
        <p>Página criada para a próxima fase. Aqui poderá entrar controle de AWB, coleta, XML, CTE, embarque por filial e integração com o fluxo de rastreamento.</p>
    </section>
    """
    return page("Expedição", body, user)


def render_usuarios(user, msg=""):
    conn = connect()
    users = conn.execute("""
        SELECT u.*, g.nome AS grupo_nome
        FROM users u LEFT JOIN grupos g ON g.id=u.grupo_id
        ORDER BY u.perfil, u.login
    """).fetchall()
    linhas = []
    for u in users:
        linhas.append(f"""
        <tr>
            <td>{html_escape(u['login'])}</td><td>{html_escape(u['nome'])}</td><td>{html_escape(u['perfil'])}</td>
            <td>{html_escape(u['grupo_nome'])}</td><td>{html_escape(u['cliente_codigo'])}</td>
            <td>
                <form method='post' action='/resetar_senha' class='inline'>
                    <input type='hidden' name='login' value='{html_escape(u['login'])}'>
                    <input name='senha' placeholder='Nova senha' required minlength='6'>
                    <button>Resetar</button>
                </form>
            </td>
        </tr>
        """)
    conn.close()
    body = f"""
    {'<div class="okmsg">'+html_escape(msg)+'</div>' if msg else ''}
    <section class='card'>
        <h2>Criar usuário</h2>
        <form method='post' action='/criar_usuario' class='user_form'>
            <div><label>Login</label><input name='login' required></div>
            <div><label>Nome</label><input name='nome' required></div>
            <div><label>Perfil</label><select name='perfil'>{perfil_options()}</select></div>
            <div><label>Grupo</label><select name='grupo_id'>{select_options('grupos','id','nome')}</select></div>
            <div><label>Código cliente</label><input name='cliente_codigo'></div>
            <div><label>Senha</label><input name='senha' minlength='6' required></div>
            <button>Criar</button>
        </form>
    </section>
    <section class='card'><h2>Usuários cadastrados</h2><div class='tablebox'><table><tr><th>Login</th><th>Nome</th><th>Perfil</th><th>Grupo</th><th>Cliente</th><th>Reset</th></tr>{''.join(linhas)}</table></div></section>
    """
    return page("Usuários", body, user)


def render_grupos(user, msg=""):
    conn = connect()
    grupos = conn.execute("""
        SELECT g.*,
               (SELECT COUNT(*) FROM demandas d WHERE d.active=1 AND d.finalizado=0 AND d.grupo_id=g.id) AS qtd_demandas,
               (SELECT COUNT(DISTINCT parceiro) FROM demandas d WHERE d.active=1 AND d.finalizado=0 AND d.grupo_id=g.id) AS qtd_parceiros
        FROM grupos g
        ORDER BY g.nome
    """).fetchall()
    linhas = []
    for g in grupos:
        pessoas = [x for x in (norm_text(g['pessoa1_login']), norm_text(g['pessoa2_login'])) if x]
        pessoa_buttons = ""
        for pessoa in pessoas:
            pessoa_buttons += f"<a class='btn tiny' href='/download_demanda_grupo?grupo_id={g['id']}&pessoa={html_escape(pessoa)}'>Baixar {html_escape(pessoa)}</a>"
        linhas.append(f"""
        <tr>
            <td>{g['id']}</td>
            <td>{html_escape(g['nome'])}</td>
            <td>
                <form method='post' action='/atualizar_grupo' class='inline group_edit'>
                    <input type='hidden' name='grupo_id' value='{g['id']}'>
                    <input name='nome' value='{html_escape(g['nome'])}' placeholder='Grupo'>
                    <input name='pessoa1_login' value='{html_escape(g['pessoa1_login'])}' placeholder='Pessoa 1 login'>
                    <input name='pessoa2_login' value='{html_escape(g['pessoa2_login'])}' placeholder='Pessoa 2 login'>
                    <input name='max_diario' type='number' min='0' value='{g['max_diario'] or 0}' placeholder='Máx.'>
                    <button>Salvar</button>
                </form>
            </td>
            <td>{g['qtd_demandas'] or 0}</td>
            <td>{g['qtd_parceiros'] or 0}</td>
            <td>{'Ativo' if g['ativo'] else 'Inativo'}</td>
            <td class='actions_cell'>
                <a class='btn small' href='/download_demanda_grupo?grupo_id={g['id']}'>Baixar grupo</a>
                {pessoa_buttons}
            </td>
        </tr>
        """)
    conn.close()
    body = f"""
    {'<div class="okmsg">'+html_escape(msg)+'</div>' if msg else ''}
    <section class='card'>
        <h2>Criar grupo</h2>
        <p class='note'>Depois de criar ou alterar as pessoas, clique em <b>Gerar/Redistribuir demanda</b> para vincular somente demandas em aberto aos grupos e às pessoas.</p>
        <form method='post' action='/criar_grupo' class='user_form'>
            <div><label>Nome do grupo</label><input name='nome' placeholder='Grupo 1' required></div>
            <div><label>Pessoa 1 login</label><input name='pessoa1_login' placeholder='emilia'></div>
            <div><label>Pessoa 2 login</label><input name='pessoa2_login' placeholder='tayane'></div>
            <div><label>Máximo diário</label><input name='max_diario' type='number' min='0' value='0'></div>
            <button>Criar grupo</button>
        </form>
    </section>
    <section class='card'>
        <h2>Gerar demanda para grupos/pessoas</h2>
        <p>Distribui somente as demandas em aberto da base ativa por parceiro/base da coluna AC. Fechados ficam fora da fila operacional.</p>
        <form method='post' action='/redistribuir_demanda' class='actions left'>
            <button>Gerar/Redistribuir demanda agora</button>
            <a class='btn secondary' href='/rastreamento'>Ver dashboard</a>
        </form>
    </section>
    <section class='card'>
        <h2>Grupos cadastrados</h2>
        <div class='tablebox'>
            <table>
                <tr><th>ID</th><th>Grupo</th><th>Adicionar/alterar pessoas</th><th>Demandas em aberto</th><th>Parceiros</th><th>Status</th><th>Downloads</th></tr>
                {''.join(linhas)}
            </table>
        </div>
    </section>
    """
    return page("Grupos", body, user)

def render_pesquisa_nf(user, nf):
    rows = fetch_demands(user, {"nf": nf}, limit=100)
    body = f"""
    <section class='card'>
        <form method='get' action='/pesquisa_nf' class='search_form'>
            <input name='nf' value='{html_escape(nf)}' placeholder='Digite a NF'>
            <button>Pesquisar</button>
        </form>
    </section>
    <section class='card'><h2>Resultado da pesquisa por NF</h2>{table_demands(rows)}</section>
    """
    return page("Pesquisa por NF", body, user)


def create_group_workbook(rows, title="Demanda do Grupo"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Demanda"
    headers = [
        "NF", "Status Atualizacao", "Ocorrencia Atualizacao", "Responsavel",
        "Status Sistema", "Grupo", "Parceiro", "UF", "Cliente", "Data Prevista",
        "Dias Atraso SLA", "Atraso Atualizacao Dias", "Cidade Destino", "Filial", "Tipo"
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("nf"), "", "", r.get("responsavel_login") or "",
            r.get("status_sistema"), r.get("grupo_nome"), r.get("parceiro"), r.get("estado"),
            r.get("cliente_nome"), fmt_date(r.get("data_prevista")), r.get("dias_atraso_sla") or 0,
            r.get("atraso_atualizacao_dias") or 0, r.get("cidade_destino"), r.get("filial"), r.get("tipo")
        ])

    fill = PatternFill("solid", fgColor="0B6B8F")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for cell_obj in ws[1]:
        cell_obj.fill = fill
        cell_obj.font = font
        cell_obj.alignment = Alignment(horizontal="center")
        cell_obj.border = Border(bottom=thin)

    widths = [18, 26, 36, 18, 24, 18, 28, 10, 42, 16, 16, 22, 24, 24, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"B2:B{ws.max_row}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def rows_to_csv(rows):
    out = io.StringIO()
    w = csv.writer(out, delimiter=CSV_SEPARATOR)
    w.writerow(["NF","Status","Grupo","Parceiro","UF","Cliente","Data Prevista","Atraso SLA","Atraso Atualizacao","Responsavel","Ocorrencia"])
    for r in rows:
        w.writerow([r.get("nf"), r.get("status_grupo") or r.get("status_sistema"), r.get("grupo_nome"), r.get("parceiro"), r.get("estado"), r.get("cliente_nome"), fmt_date(r.get("data_prevista")), r.get("dias_atraso_sla"), r.get("atraso_atualizacao_dias"), r.get("responsavel_login"), r.get("ocorrencia_grupo") or r.get("ocorrencia_original")])
    return out.getvalue().encode("utf-8-sig")


CSS = r"""
<style>
*{box-sizing:border-box}body{margin:0;background:#edf3f8;font-family:Segoe UI,Arial,sans-serif;color:#152536}.wrap{max-width:1180px;margin:24px auto;padding:0 16px}.wide{max-width:1500px}
.top{display:flex;justify-content:space-between;align-items:flex-start;gap:14px;margin-bottom:18px}.top h1{margin:0;font-size:28px}.top p{margin:6px 0 0;color:#607489}.nav{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}.nav a,.btn,button{background:#0b6b8f;color:#fff;border:0;border-radius:12px;padding:10px 15px;text-decoration:none;font-weight:800;cursor:pointer}.btn.secondary,.nav a:nth-last-child(2){background:#607d8b}.actions{display:flex;gap:10px;flex-wrap:wrap}.actions.left{justify-content:flex-start}.card{background:white;border:1px solid #dfeaf2;border-radius:18px;padding:20px;box-shadow:0 10px 26px rgba(31,63,94,.07);margin-bottom:16px}.narrow{max-width:520px}
.grid{display:grid;gap:14px;margin-bottom:16px}.two{grid-template-columns:1fr 1fr}.four{grid-template-columns:repeat(4,1fr)}.five{grid-template-columns:repeat(5,1fr)}.six{grid-template-columns:repeat(6,1fr)}.kpis{grid-template-columns:repeat(5,1fr)}
.kpi{border-radius:18px;color:white;padding:18px;min-height:118px;box-shadow:0 8px 20px rgba(31,63,94,.11)}.kpi b{font-size:30px;display:block}.kpi span{display:block;font-weight:900}.kpi small{opacity:.95}.blue{background:#159ad1}.green{background:#00a957}.orange{background:#f06b2f}.red{background:#c94343}.purple{background:#8257d8}.gray{background:#697586}
.module{display:block;padding:22px;border-radius:18px;background:white;color:#163247;text-decoration:none;font-size:20px;font-weight:900;border:1px solid #dfeaf2;box-shadow:0 8px 20px rgba(31,63,94,.07)}.module small{font-size:12px;color:#607489}.module.active{background:linear-gradient(135deg,#0b6b8f,#10b36a);color:white}.module.active small{color:white}
.form,.upload{display:flex;flex-direction:column;gap:9px}.form input,.form select,.upload input,.user_form input,.user_form select,.filter_form input,.filter_form select,.search_form input{padding:12px;border:1px solid #cbdce8;border-radius:12px;font-size:14px}.upload{gap:12px}.upload button{align-self:flex-start}
.user_form{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;align-items:end}.user_form div,.filter_form div{display:flex;flex-direction:column;gap:6px}.filter_form{display:grid;grid-template-columns:repeat(4,1fr) auto auto;gap:12px;align-items:end}.filtros_sla_card{overflow:hidden;padding:24px 22px}.filter_form_sla{grid-template-columns:170px 150px 120px minmax(260px,1.05fr) minmax(160px,.65fr) minmax(210px,.8fr) minmax(170px,.7fr) 236px;gap:12px;align-items:end}.filtros_sla_title{justify-content:flex-end;align-self:end;padding-bottom:4px}.filtros_sla_title h2{margin:0;font-size:26px;white-space:nowrap}.filter_form_sla .campo_mes select{width:150px}.filter_form_sla .campo_ano input{width:120px}.filter_form_sla .campo_cliente select{width:100%;max-width:300px}.filter_form_sla .campo_grupo select{width:100%}.filter_form_sla .campo_parceiro select{width:100%}.filter_form_sla .campo_estado select{width:100%}.filter_actions{display:grid!important;grid-template-columns:1fr 1fr;gap:10px;align-items:end}.filter_actions .btn,.filter_actions button{height:40px;min-width:105px;text-align:center;display:flex;align-items:center;justify-content:center}.search_form{display:flex;gap:10px}.search_form input{min-width:280px}
.tablebox{overflow:auto;max-height:620px;border:1px solid #edf2f7;border-radius:12px}.tablebox.small{max-height:360px}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px 10px;border-bottom:1px solid #e7eef5;white-space:nowrap;text-align:left}th{background:#e7f3ec;color:#102b1d;position:sticky;top:0}tr:nth-child(even) td{background:#f7fbff}
.barline{margin:14px 0}.barline>div:first-child{display:flex;justify-content:space-between}.bar{height:14px;background:#edf2f6;border-radius:20px;overflow:hidden;margin-top:7px}.bar i{display:block;height:100%;background:#0b6b8f}
.badge{display:inline-block;border-radius:999px;padding:5px 9px;font-weight:900}.badge.bad{background:#ffe1e1;color:#9b1c1c}.badge.warn{background:#fff2c2;color:#775000}.badge.ok{background:#dffbe8;color:#0b6b39}
.login_wrap{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px;background:linear-gradient(135deg,#eaf4fb,#f3fbf6)}.login_card{width:100%;max-width:430px;background:#fff;border-radius:22px;padding:28px;box-shadow:0 16px 38px rgba(31,63,94,.14)}.brand{background:linear-gradient(135deg,#0b6b8f,#10b36a);color:#fff;border-radius:18px;padding:22px;margin-bottom:18px}.brand h1{margin:0}.brand p{margin:6px 0 0}.note{color:#65788a}.err{background:#ffe8e8;color:#8d1f1f;border:1px solid #ffc4c4;padding:12px;border-radius:12px;margin-bottom:12px}.okmsg{background:#e8fff1;color:#146c38;border:1px solid #b8efcb;padding:12px;border-radius:12px;margin-bottom:12px}.inline{display:flex;gap:6px;flex-wrap:wrap}.inline input{padding:7px;border:1px solid #cbdce8;border-radius:8px}.group_edit input{max-width:155px}.btn.tiny{padding:6px 8px;font-size:11px;margin:2px}.actions_cell{display:flex;gap:5px;flex-wrap:wrap}
@media(max-width:1100px){.two,.four,.five,.six,.kpis,.filter_form,.filter_form_sla,.user_form{grid-template-columns:1fr}.top{display:block}.nav{justify-content:flex-start;margin-top:12px}.filtros_sla_title{justify-content:flex-start}.filter_actions{display:flex!important}}
</style>
"""


class App(BaseHTTPRequestHandler):
    def send_html(self, body, code=200):
        data = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_file(self, data, filename, content_type):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, location):
        self.send_response(302)
        self.send_header("Location", location)
        self.end_headers()

    def require_user(self):
        user = get_session_user(self)
        if not user:
            self.send_html(render_login("Faça login para acessar o sistema."), 401)
            return None
        return user

    def do_GET(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)

        if parsed.path == "/login":
            self.send_html(render_login())
            return

        if parsed.path == "/logout":
            self.send_response(302)
            clear_session(self)
            self.send_header("Location", "/login")
            self.end_headers()
            return

        user = self.require_user()
        if not user:
            return

        if user.get("trocar_senha") and parsed.path not in ("/alterar_senha",):
            self.send_html(render_change_password(user, "Altere sua senha inicial para continuar."))
            return

        if parsed.path == "/alterar_senha":
            self.send_html(render_change_password(user))
            return

        if parsed.path == "/":
            if not has_permission(user, ("admin",)):
                self.redirect("/rastreamento")
                return
            self.send_html(render_admin(user))
            return

        if parsed.path == "/gestor":
            self.redirect("/acompanhamento")
            return

        if parsed.path == "/acompanhamento":
            if not has_permission(user, ("admin", "gestor", "rastreamento")):
                self.send_html(page("Acesso negado", "<div class='err'>Acesso restrito ao acompanhamento.</div>", user), 403)
                return
            self.send_html(render_acompanhamento(user, qs))
            return

        if parsed.path == "/rastreamento":
            if not has_permission(user, ("admin", "gestor", "rastreamento")):
                self.send_html(page("Acesso negado", "<div class='err'>Acesso restrito ao rastreamento.</div>", user), 403)
                return
            self.send_html(render_rastreamento(user, qs))
            return

        if parsed.path == "/chamado_parceiro":
            if not has_permission(user, ("admin", "gestor", "chamado_parceiro")):
                self.send_html(page("Acesso negado", "<div class='err'>Acesso restrito ao Chamado Parceiro.</div>", user), 403)
                return
            self.send_html(render_chamado_parceiro(user, qs))
            return

        if parsed.path == "/cliente":
            if not has_permission(user, ("admin", "gestor", "cliente")):
                self.send_html(page("Acesso negado", "<div class='err'>Acesso restrito ao cliente.</div>", user), 403)
                return
            self.send_html(render_cliente(user, qs))
            return

        if parsed.path == "/expedicao":
            if not has_permission(user, ("admin", "gestor", "expedicao")):
                self.send_html(page("Acesso negado", "<div class='err'>Acesso restrito à expedição.</div>", user), 403)
                return
            self.send_html(render_expedicao(user, qs))
            return

        if parsed.path == "/usuarios":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            self.send_html(render_usuarios(user))
            return

        if parsed.path == "/grupos":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            self.send_html(render_grupos(user))
            return

        if parsed.path == "/pesquisa_nf":
            nf = qs.get("nf", [""])[0]
            self.send_html(render_pesquisa_nf(user, nf))
            return

        if parsed.path == "/download_demanda_grupo":
            if not has_permission(user, ("admin", "gestor", "rastreamento")):
                self.send_html(page("Acesso negado", "<div class='err'>Sem permissão para baixar demanda de grupo.</div>", user), 403)
                return
            grupo_id = qs.get("grupo_id", [""])[0]
            pessoa = norm_text(qs.get("pessoa", [""])[0])
            if user.get("perfil") == "rastreamento":
                grupo_id = str(user.get("grupo_id") or "")
            if not grupo_id:
                self.send_html(page("Grupo obrigatório", "<div class='err'>Informe o grupo para baixar a demanda.</div>", user), 400)
                return
            rows = fetch_demands(user, {"grupo_id": grupo_id}, limit=MAX_QUERY_ROWS)
            rows = [r for r in rows if not r.get("finalizado")]
            if pessoa:
                rows = [r for r in rows if norm_key(r.get("responsavel_login")) == norm_key(pessoa)]
            data = create_group_workbook(rows, "Demanda")
            sufixo = f"_{safe_filename(pessoa)}" if pessoa else ""
            self.send_file(data, f"demanda_grupo_{safe_filename(grupo_id)}{sufixo}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return

        if parsed.path == "/download_grupo":
            if user.get("perfil") != "rastreamento":
                self.send_html(page("Acesso negado", "<div class='err'>Somente usuários do rastreamento.</div>", user), 403)
                return
            modo = qs.get("modo", ["meu"])[0]
            filters = {"grupo_id": str(user.get("grupo_id") or "")}
            rows = fetch_demands(user, filters, limit=MAX_QUERY_ROWS)
            rows = [r for r in rows if not r.get("finalizado")]
            if modo == "meu":
                # Para não travar o operacional, se não houver responsável definido, baixa a demanda do grupo.
                own = [r for r in rows if norm_key(r.get("responsavel_login")) == norm_key(user.get("login"))]
                if own:
                    rows = own
            data = create_group_workbook(rows, "Demanda")
            self.send_file(data, f"demanda_{modo}_{safe_filename(user.get('login'))}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return

        if parsed.path == "/download_relatorio_parceiro":
            if not has_permission(user, ("admin", "gestor", "chamado_parceiro")):
                self.send_html(page("Acesso negado", "<div class='err'>Sem permissão para relatórios de parceiro.</div>", user), 403)
                return
            tipo = qs.get("tipo", [""])[0]
            filters = parceiro_filters_from_qs(qs)
            rows = parceiro_report_rows(user, filters, tipo)
            title = parceiro_report_title(tipo)
            data = create_partner_report_workbook(rows, title)
            filename = f"relatorio_parceiro_{safe_filename(tipo)}_{safe_filename(title)}.xlsx"
            self.send_file(data, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return

        if parsed.path == "/download_todos":
            if not has_permission(user, ("admin", "gestor")):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin/gestor.</div>", user), 403)
                return
            rows = fetch_demands(user, {}, limit=MAX_QUERY_ROWS)
            self.send_file(rows_to_csv(rows), "base_ativa_rastreamento.csv", "text/csv; charset=utf-8")
            return

        self.send_html(page("Página não encontrada", "<div class='err'>Rota não encontrada.</div>", user), 404)

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path == "/login":
            form = read_form_urlencoded(self)
            login = norm_text(form.get("login"))
            senha = form.get("senha", "")
            conn = connect()
            u = conn.execute("SELECT * FROM users WHERE login=? AND ativo=1", (login,)).fetchone()
            conn.close()
            if not u or not check_password(senha, u["senha_hash"]):
                self.send_html(render_login("Usuário ou senha inválidos."), 401)
                return
            self.send_response(302)
            create_session(self, login)
            perfil = u["perfil"]
            dest = "/" if perfil == "admin" else "/acompanhamento" if perfil == "gestor" else f"/{perfil}" if perfil in ("cliente", "expedicao") else "/rastreamento"
            self.send_header("Location", dest)
            self.end_headers()
            return

        user = self.require_user()
        if not user:
            return

        if parsed.path == "/alterar_senha":
            form = read_form_urlencoded(self)
            atual = form.get("senha_atual", "")
            nova = form.get("nova_senha", "")
            confirmar = form.get("confirmar_senha", "")
            conn = connect()
            u = conn.execute("SELECT * FROM users WHERE login=?", (user["login"],)).fetchone()
            if not check_password(atual, u["senha_hash"]):
                conn.close()
                self.send_html(render_change_password(user, "Senha atual incorreta."), 400)
                return
            if len(nova) < 6 or nova != confirmar:
                conn.close()
                self.send_html(render_change_password(user, "Nova senha inválida ou confirmação diferente."), 400)
                return
            conn.execute("UPDATE users SET senha_hash=?, trocar_senha=0 WHERE login=?", (hash_password(nova), user["login"]))
            conn.commit()
            conn.close()
            self.redirect("/")
            return

        if parsed.path == "/upload_geral":
            if not has_permission(user, ("admin", "gestor")):
                self.send_html(page("Acesso negado", "<div class='err'>Somente Admin/Gestor pode subir relatório geral.</div>", user), 403)
                return
            try:
                fields, files = parse_multipart(self)
                f = files.get("file")
                if not f:
                    raise ValueError("Nenhum arquivo recebido.")
                filename = safe_filename(f["filename"])
                path = os.path.join(UPLOAD_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}")
                with open(path, "wb") as out:
                    out.write(f["content"])
                ano_controle = to_int(fields.get("ano_controle"), date.today().year)
                mes_controle = to_int(fields.get("mes_controle"), date.today().month)
                total, upload_id = import_master_report(path, filename, user["login"], ano_controle, mes_controle)
                msg = f"Relatório geral importado com sucesso. Mês/Ano: {mes_controle:02d}/{ano_controle}. Demandas ativas: {total}. Upload ID: {upload_id}."
                self.send_html(render_admin(user, msg) if user["perfil"] == "admin" else render_gestor(user, {}))
            except Exception as e:
                traceback.print_exc()
                self.send_html(page("Erro no upload", f"<div class='err'>Erro ao importar: {html_escape(e)}</div>", user), 500)
            return

        if parsed.path == "/upload_retorno":
            if not has_permission(user, ("rastreamento", "admin", "gestor")):
                self.send_html(page("Acesso negado", "<div class='err'>Sem permissão para retorno.</div>", user), 403)
                return
            try:
                fields, files = parse_multipart(self)
                f = files.get("file")
                if not f:
                    raise ValueError("Nenhum arquivo recebido.")
                filename = safe_filename(f["filename"])
                path = os.path.join(UPLOAD_DIR, f"retorno_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}")
                with open(path, "wb") as out:
                    out.write(f["content"])
                updated, upload_id = update_group_report(path, filename, user["login"])
                self.send_html(page("Retorno atualizado", f"<div class='okmsg'>Retorno processado. Linhas atualizadas: {updated}. Upload ID: {upload_id}.</div><a class='btn' href='/rastreamento'>Voltar</a>", user))
            except Exception as e:
                traceback.print_exc()
                self.send_html(page("Erro no retorno", f"<div class='err'>Erro ao atualizar retorno: {html_escape(e)}</div>", user), 500)
            return

        if parsed.path == "/redistribuir_demanda":
            if not has_permission(user, ("admin", "gestor")):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin/gestor pode redistribuir demanda.</div>", user), 403)
                return
            try:
                demandas, parceiros, responsaveis = redistribuir_demandas_ativas(force=True)
                msg = f"Demanda em aberto gerada com sucesso. Demandas alocadas: {demandas}. Parceiros distribuídos: {parceiros}. Responsáveis aplicados: {responsaveis}."
                self.send_html(render_grupos(user, msg) if user.get("perfil") == "admin" else page("Redistribuição concluída", f"<div class='okmsg'>{html_escape(msg)}</div><a class='btn' href='/rastreamento'>Ver dashboard</a>", user))
            except Exception as e:
                traceback.print_exc()
                self.send_html(page("Erro", f"<div class='err'>Erro ao redistribuir: {html_escape(e)}</div>", user), 500)
            return

        if parsed.path == "/atualizar_grupo":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            form = read_form_urlencoded(self)
            grupo_id = to_int(form.get("grupo_id"), 0)
            nome = norm_text(form.get("nome"))
            p1 = norm_text(form.get("pessoa1_login"))
            p2 = norm_text(form.get("pessoa2_login"))
            maxd = to_int(form.get("max_diario"), 0)
            conn = connect()
            try:
                conn.execute("UPDATE grupos SET nome=?, pessoa1_login=?, pessoa2_login=?, max_diario=? WHERE id=?", (nome, p1, p2, maxd, grupo_id))
                for login in (p1, p2):
                    if login:
                        conn.execute("UPDATE users SET grupo_id=? WHERE login=?", (grupo_id, login))
                alternar_responsaveis_grupo(conn, grupo_id)
                conn.commit()
                self.send_html(render_grupos(user, "Grupo atualizado e pessoas vinculadas somente às demandas em aberto."))
            except Exception as e:
                conn.rollback()
                self.send_html(render_grupos(user, f"Erro: {e}"), 400)
            finally:
                conn.close()
            return

        if parsed.path == "/criar_grupo":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            form = read_form_urlencoded(self)
            nome = norm_text(form.get("nome"))
            p1 = norm_text(form.get("pessoa1_login"))
            p2 = norm_text(form.get("pessoa2_login"))
            maxd = to_int(form.get("max_diario"), 0)
            conn = connect()
            try:
                conn.execute("""
                    INSERT INTO grupos(nome, pessoa1_login, pessoa2_login, max_diario, ativo, created_at)
                    VALUES(?,?,?,?,1,?)
                """, (nome, p1, p2, maxd, datetime.now().isoformat()))
                gid = conn.execute("SELECT id FROM grupos WHERE nome=?", (nome,)).fetchone()["id"]
                for login in (p1, p2):
                    if login:
                        conn.execute("UPDATE users SET grupo_id=? WHERE login=?", (gid, login))
                conn.commit()
                try:
                    demandas, parceiros, responsaveis = redistribuir_demandas_ativas(force=True)
                    msg = f"Grupo criado com sucesso. Demanda em aberto gerada: {demandas} linhas, {parceiros} parceiros, {responsaveis} responsáveis."
                except Exception:
                    traceback.print_exc()
                    msg = "Grupo criado com sucesso. Clique em Gerar/Redistribuir demanda para alocar as demandas em aberto."
                self.send_html(render_grupos(user, msg))
            except Exception as e:
                conn.rollback()
                self.send_html(render_grupos(user, f"Erro: {e}"), 400)
            finally:
                conn.close()
            return

        if parsed.path == "/criar_usuario":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            form = read_form_urlencoded(self)
            login = norm_text(form.get("login"))
            nome = norm_text(form.get("nome")) or login
            perfil = norm_text(form.get("perfil"))
            senha = form.get("senha", "")
            grupo_id = to_int(form.get("grupo_id"), 0) or None
            cliente_codigo = norm_text(form.get("cliente_codigo"))
            if perfil == "cliente" and not cliente_codigo:
                self.send_html(render_usuarios(user, "Cliente precisa de código cliente."), 400)
                return
            conn = connect()
            try:
                conn.execute("""
                    INSERT INTO users(login, senha_hash, nome, perfil, grupo_id, cliente_codigo, ativo, trocar_senha, created_at)
                    VALUES(?,?,?,?,?,?,?,?,?)
                """, (login, hash_password(senha), nome, perfil, grupo_id, cliente_codigo, 1, 1, datetime.now().isoformat()))
                conn.commit()
                self.send_html(render_usuarios(user, "Usuário criado com sucesso."))
            except Exception as e:
                conn.rollback()
                self.send_html(render_usuarios(user, f"Erro: {e}"), 400)
            finally:
                conn.close()
            return

        if parsed.path == "/resetar_senha":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin.</div>", user), 403)
                return
            form = read_form_urlencoded(self)
            login = norm_text(form.get("login"))
            senha = form.get("senha", "")
            conn = connect()
            conn.execute("UPDATE users SET senha_hash=?, trocar_senha=1 WHERE login=?", (hash_password(senha), login))
            conn.commit()
            conn.close()
            self.send_html(render_usuarios(user, f"Senha de {login} resetada."))
            return

        if parsed.path == "/criar_perfil":
            if not has_permission(user, ("admin",)):
                self.send_html(page("Acesso negado", "<div class='err'>Somente admin pode criar perfil.</div>", user), 403)
                return
            form = read_form_urlencoded(self)
            nome = norm_text(form.get("nome"))
            descricao = norm_text(form.get("descricao"))
            lista = []
            for codigo, _label in PERMISSOES_SISTEMA:
                if form.get("perm_" + codigo):
                    lista.append(codigo)
            if not nome:
                self.send_html(render_admin(user, "Informe o nome do perfil."), 400)
                return
            conn = connect()
            try:
                conn.execute("INSERT OR IGNORE INTO perfis(nome, descricao, ativo, created_at) VALUES(?,?,1,?)", (nome, descricao, datetime.now().isoformat()))
                conn.execute("UPDATE perfis SET descricao=?, ativo=1 WHERE nome=?", (descricao, nome))
                conn.execute("DELETE FROM perfil_permissoes WHERE perfil_nome=?", (nome,))
                for permissao in lista:
                    conn.execute("INSERT OR IGNORE INTO perfil_permissoes(perfil_nome, permissao) VALUES(?,?)", (nome, permissao))
                conn.commit()
                self.send_html(render_admin(user, "Perfil criado/atualizado."))
            except Exception as e:
                conn.rollback()
                self.send_html(render_admin(user, f"Erro ao salvar perfil: {e}"), 400)
            finally:
                conn.close()
            return

        self.send_html(page("Rota inválida", "<div class='err'>Rota POST inválida.</div>", user), 404)


if __name__ == "__main__":
    init_db()
    load_sessions()
    print("=" * 70)
    print("SISTEMA DE RASTREAMENTO POR GRUPOS - PRO INICIAL")
    print(f"Banco de dados: {DB_PATH}")
    print(f"Uploads: {UPLOAD_DIR}")
    print(f"Acesse: http://127.0.0.1:{PORT}")
    print(f"Login inicial: {ADMIN_LOGIN}")
    print("Senha inicial: Admin@123")
    print("=" * 70)
    HTTPServer((HOST, PORT), App).serve_forever()

